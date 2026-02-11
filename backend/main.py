# Datei: main.py

from fastapi import FastAPI, UploadFile, HTTPException, File, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
import base64
import os
import json
import fitz  # PyMuPDF
from openai import OpenAI
from dotenv import load_dotenv
from io import BytesIO
import numpy as np
import re
# Note: Plotly removed in Phase 2.2 - using Chart.js in frontend (lighter weight)
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, Alignment
from pydantic import BaseModel
import prompts  # Import comprehensive Socratic prompts
import session_manager  # Session management for Self/Business Clarity
import clarity_endpoints  # Self Clarity and Business Clarity endpoints

# 🔹 Umgebung laden (.env mit OPENAI_API_KEY)
load_dotenv()

app = FastAPI()

# Setup Self Clarity and Business Clarity routes
clarity_endpoints.setup_clarity_routes(app)
client = OpenAI(
    api_key=os.getenv("OPENAI_API_KEY"),
    timeout=60.0,  # 60 second timeout to prevent hanging forever
    max_retries=2
)

# 🔹 CORS freischalten (Frontend darf auf Backend zugreifen)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In Produktion: besser gezielt Domain angeben
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ------------------------------
# Favicon
# ------------------------------
@app.get("/favicon.ico", include_in_schema=False)
async def favicon():
    return FileResponse("favicon.ico")

# ------------------------------
# Test-Endpoint
# ------------------------------
@app.get("/")
def read_root():
    return {"message": "Clarity Coach API läuft ✅"}

# ------------------------------
# Health Check
# ------------------------------
@app.get("/health")
def health_check():
    """Check if backend and OpenAI API are working"""
    api_key = os.getenv("OPENAI_API_KEY")
    
    if not api_key:
        return {"status": "error", "message": "OpenAI API key not found"}
    
    if api_key == "your-api-key-here":
        return {"status": "error", "message": "OpenAI API key not configured (still using placeholder)"}
    
    # Test OpenAI connection
    try:
        test_response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": "test"}],
            max_tokens=5
        )
        return {
            "status": "ok", 
            "message": "Backend and OpenAI API working",
            "api_key_valid": True
        }
    except Exception as e:
        return {
            "status": "error",
            "message": f"OpenAI API error: {str(e)}",
            "api_key_valid": False
        }


# ------------------------------
# Hilfsfunktion: Clarity-Coach-Prompt ausführen
# ------------------------------
def run_clarity_coach(full_text: str):
    """
    Nimmt reinen Aufgabentext und gibt die strukturierte Aufgabenliste zurück:
    [
      {
        "number": "1",
        "topic": "...",
        "difficulty": "leicht/mittel/anspruchsvoll",
        "task": "volle Aufgabenangabe",
        "subtasks": [
          {
            "label": "a",
            "task": "Text der Teilaufgabe a",
            "questions": ["...", "..."]
          }
        ]
      },
      ...
    ]
    """

    clarity_prompt = f"""
Du bist der KI-Entwicklungsassistent für das Projekt Clarity Coach.

Deine Aufgabe:
- Analysiere den folgenden Aufgabentext (mehrere Aufgaben mit Teilaufgaben möglich).
- Erkenne Aufgaben (1., 2., 3., …) und Teilaufgaben (a), b), c), …).
- Erstelle für jede Aufgabe und jede Teilaufgabe:
  • "number": Aufgabennummer als String (z.B. "1")
  • "topic": kurzes Thema (z.B. "Kubische Gleichungen", "Potenzfunktionen")
  • "difficulty": "leicht", "mittel" oder "anspruchsvoll"
  • "task": vollständiger Text der übergeordneten Aufgabe (ohne die einzelnen Teilaufgaben)
  • "subtasks": Liste von Objekten mit:
      - "label": Buchstabe der Teilaufgabe, z.B. "a"
      - "task": Text der Teilaufgabe
      - "questions": 3–5 sokratische Fragen (Strings)

WICHTIG: Die Fragen müssen **sehr aufgabenspezifisch** sein und sich konkret auf Terme, Zahlen und Begriffe der jeweiligen Teilaufgabe beziehen.

Für jede Teilaufgabe erstelle Fragen aus diesen Kategorien (in beliebiger Reihenfolge):

1. STRUKTUR-FRAGE
   - Frage nach der Form der Gleichung/Funktion und ihren Bestandteilen.
   - Nenne mindestens einen konkreten Term oder eine Zahl aus der Aufgabe.
   - Beispiel: „Welche Zahl wird in der Gleichung x^3 - 27 = 0 als Kubikzahl verwendet?“

2. UMFORMUNGS-FRAGE
   - Frage nach einem konkreten nächsten Rechenschritt.
   - Beispiel: „Wie kannst du die -27 in der Gleichung x^3 - 27 = 0 auf die andere Seite bringen?“

3. OPERATIONS-FRAGE
   - Frage nach der passenden Rechenoperation/Funktion (z.B. Wurzel, Logarithmus, Ableitung).
   - Beispiel: „Welche Umkehrfunktion brauchst du, um aus x^3 wieder x zu erhalten?“

4. KONTROLL- ODER INTERPRETATIONS-FRAGE
   - Frage nach der Bedeutung des Ergebnisses, der Anzahl der Lösungen, Monotonie, etc.
   - Beispiel: „Was sagt dir die Tatsache, dass f'(x) = 3x^2 ≥ 0 für alle x über die Anzahl der Nullstellen von f(x) = x^3 - 27?“

Regeln für die Formulierung:
- Jede Frage muss mindestens EIN konkretes Element aus der Teilaufgabe enthalten
  (z.B. eine Zahl wie 27 oder 1/27, einen Term wie x^3, f(x), oder einen Fachbegriff wie „Nullstelle“).
- Vermeide generische Fragen wie:
  • „Was sagt dir die Gleichung über x?“
  • „Wie kannst du die Gleichung lösen?“
  • „Welche Schritte musst du machen?“
- Formuliere die Fragen so, dass der Schüler *konkrete* nächste Schritte beschreiben muss
  (z.B. „Welche Zahl…“, „Welchen Term…“, „Welche Umformung…“, „Welche Eigenschaft…“).

Ausgabeformat:
- Gib ein JSON-OBJEKT mit GENAU einem Feld "tasks" zurück.
- "tasks" ist eine LISTE von Aufgabenobjekten wie im folgenden Schema:

{{
  "tasks": [
    {{
      "number": "1",
      "topic": "Kubische Gleichungen",
      "difficulty": "mittel",
      "task": "Zeige, dass die Gleichung nur eine reelle Lösung besitzt.",
      "subtasks": [
        {{
          "label": "a",
          "task": "x^3 - 27 = 0",
          "questions": [
            "Frage 1 ...",
            "Frage 2 ...",
            "Frage 3 ..."
          ]
        }}
      ]
    }}
  ]
}}

Keine Erklärtexte außerhalb dieses JSON-Objekts, keine Markdown-Codeblöcke.

Hier ist der vollständige Aufgabentext:

{full_text}
"""

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        response_format={"type": "json_object"},
        messages=[
            {
                "role": "system",
                "content": (
                    "Du bist ein geduldiger, sokratischer Mathematiklehrer. "
                    "Du erzeugst sehr aufgabenspezifische Fragen und gibst "
                    "die Antwort ausschließlich als gültiges JSON-Objekt mit dem Feld 'tasks' zurück."
                ),
            },
            {"role": "user", "content": clarity_prompt},
        ],
    )

    raw = response.choices[0].message.content
    
    # Safe printing for Windows console (handle Unicode characters)
    try:
        print("\n--- GPT-Raw-Response (JSON-Modus) ---\n", raw, "\n------------------------\n")
    except UnicodeEncodeError:
        print("\n--- GPT-Raw-Response (JSON-Modus) ---\n[Response contains Unicode characters - check logs]\n------------------------\n")
    
    try:
        obj = json.loads(raw)
        tasks = obj.get("tasks", None)
        if isinstance(tasks, list):
            return tasks
        if isinstance(obj, list):
            return obj
        return {
            "error": "Antwort hatte nicht das erwartete Format (Feld 'tasks' fehlt oder ist keine Liste).",
            "raw_output": raw,
        }
    except Exception as e:
        print(f"[ERROR] JSON-Fehler im JSON-Modus: {e}")
        return {
            "error": "Konnte JSON nicht korrekt verarbeiten",
            "raw_output": raw,
        }


# ------------------------------
# Textbasierte Eingabe (z.B. für Tests)
# ------------------------------
@app.post("/clarity")
def clarity(input: dict = Body(...)):
    user_input = input.get("task", "")
    if not user_input.strip():
        return {"error": "Kein Aufgabentext übergeben."}

    result = run_clarity_coach(user_input)
    return result


# ------------------------------
# REMOVED: Lösung für eine Teilaufgabe generieren
# This endpoint has been REMOVED because it directly contradicts
# the Socratic method. Showing full solutions creates learned helplessness
# and makes assessment impossible.
# 
# REPLACED BY: /hint endpoint (Progressive Hint System)
# ------------------------------


# ------------------------------
# Progressive Hint System (Socratic → Directive → Specific)
# Replaces the old /solve endpoint
# ------------------------------
@app.post("/hint")
async def get_hint(payload: dict = Body(...)):
    """
    Progressive hint system that guides students without revealing solutions.
    
    Hint Levels:
    1. Socratic: Ask guiding questions that lead to understanding
    2. Directive: Give specific steps to take without revealing answer
    3. Specific: Provide targeted help for a specific aspect
    """
    task_number = payload.get("taskNumber")
    task_text = payload.get("taskText", "")
    topic = payload.get("topic", "")
    sub_label = payload.get("subLabel")
    subtask_text = payload.get("subtaskText", "")
    hint_level = payload.get("hintLevel", 1)
    previous_hints = payload.get("previousHints", None)

    if not subtask_text or not str(subtask_text).strip():
        raise HTTPException(status_code=400, detail="Keine Teilaufgabe übergeben.")

    # Get comprehensive system prompt based on hint level
    # This uses the Math Clarity prompt optimization with question patterns
    system_prompt = prompts.get_hint_prompt(hint_level, question_pattern_type="all")

    # Build the task-specific user prompt
    hint_prompt = f"""
**Aufgabe {task_number}: {topic}**
Hauptaufgabe: {task_text}

**Teilaufgabe {sub_label}:**
{subtask_text}

{f"Vorherige Hilfestellung war: {previous_hints}" if previous_hints else ""}

Gib deine Antwort im folgenden JSON-Format:
{{
  "hint": "Dein Hinweis hier (max 2-3 Sätze)",
  "encouragement": "Ein aufmunternder Satz (z.B. 'Du bist auf dem richtigen Weg!')"
}}

REGELN:
- Halte den Hinweis kurz und prägnant
- Verwende $...$ für inline LaTeX wenn nötig
- VERRATE NICHT DIE LÖSUNG
- Gib Mut und Motivation
- Nutze die Frage-Patterns aus dem System-Prompt
"""

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": system_prompt
                },
                {"role": "user", "content": hint_prompt},
            ],
            response_format={"type": "json_object"},
            temperature=0.7
        )

        hint_json_str = response.choices[0].message.content
        
        # Safe printing
        try:
            print(f"\n--- Hint Level {hint_level} Generated ---\n{hint_json_str[:200]}...\n")
        except UnicodeEncodeError:
            print(f"\n--- Hint Level {hint_level} Generated ---\n[Contains Unicode characters]\n")

        hint_data = json.loads(hint_json_str)
        
        return {
            "hint": hint_data.get("hint", "Denke über die Grundlagen nach."),
            "encouragement": hint_data.get("encouragement", "Du schaffst das!"),
            "level": hint_level,
            "success": True
        }

    except json.JSONDecodeError as e:
        print(f"[ERROR] Hint JSON Parse Error: {e}")
        raise HTTPException(status_code=500, detail=f"Hint generation failed: {str(e)}")
    except Exception as e:
        print(f"[ERROR] Error generating hint: {e}")
        raise HTTPException(status_code=500, detail=f"Hint generation failed: {str(e)}")


# ------------------------------
# Smart Approach Checker (Phase 3.2)
# Validates student work WITHOUT revealing solution
# ------------------------------
@app.post("/check-approach")
async def check_approach(payload: dict = Body(...)):
    """
    Smart Approach Checker - Analyzes student's work and provides feedback
    WITHOUT revealing the solution.
    
    This helps students understand if they're on the right track while
    maintaining the Socratic teaching methodology.
    """
    task_number = payload.get("taskNumber")
    task_text = payload.get("taskText", "")
    topic = payload.get("topic", "")
    sub_label = payload.get("subLabel")
    subtask_text = payload.get("subtaskText", "")
    student_work = payload.get("studentWork", "")
    
    if not subtask_text or not str(subtask_text).strip():
        raise HTTPException(status_code=400, detail="Keine Teilaufgabe übergeben.")
    
    if not student_work or not str(student_work).strip():
        raise HTTPException(status_code=400, detail="Keine Schülerarbeit übergeben.")

    # Get comprehensive approach checker prompt with self-correction strategies
    system_prompt = prompts.APPROACH_CHECKER.format(
        task_number=task_number,
        topic=topic,
        task_text=task_text,
        sub_label=sub_label,
        subtask_text=subtask_text,
        student_work=student_work,
        self_correction_pattern=prompts.QUESTION_PATTERNS["self_correction"]
    )

    # Add the Math Clarity core principles
    full_system_prompt = prompts.MATH_CLARITY_CORE + "\n\n" + system_prompt

    try:
        print(f"[CHECK] Checking approach for task {task_number}{sub_label}...")

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": full_system_prompt
                },
                {"role": "user", "content": f"Analysiere die Schülerarbeit und gib Feedback gemäß den Anweisungen."},
            ],
            response_format={"type": "json_object"},
            temperature=0.7
        )

        check_json_str = response.choices[0].message.content
        
        # Safe printing
        try:
            print(f"\n--- Approach Check Result ---\n{check_json_str[:300]}...\n")
        except UnicodeEncodeError:
            print(f"\n--- Approach Check Result ---\n[Contains Unicode characters]\n")

        check_data = json.loads(check_json_str)
        
        return {
            "isOnRightTrack": check_data.get("isOnRightTrack", False),
            "overallAssessment": check_data.get("overallAssessment", "Überprüfung abgeschlossen."),
            "strengths": check_data.get("strengths", []),
            "improvements": check_data.get("improvements", []),
            "specificIssue": check_data.get("specificIssue"),
            "nextStep": check_data.get("nextStep", "Arbeite weiter an deinem Ansatz."),
            "encouragement": check_data.get("encouragement", "Du schaffst das!"),
            "confidenceScore": check_data.get("confidenceScore", 3),
            "success": True
        }

    except json.JSONDecodeError as e:
        print(f"[ERROR] Check JSON Parse Error: {e}")
        raise HTTPException(status_code=500, detail=f"Approach check failed: {str(e)}")
    except Exception as e:
        print(f"[ERROR] Error checking approach: {e}")
        raise HTTPException(status_code=500, detail=f"Approach check failed: {str(e)}")


# ------------------------------
# Datei-Upload (Bild oder PDF)
# ------------------------------
@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    try:
        print(f"[UPLOAD] Started: {file.filename}")
        
        filename = file.filename.lower()
        contents = await file.read()
        extracted_texts = []

        print(f"[UPLOAD] File size: {len(contents)} bytes")
        print(f"[UPLOAD] File type: {filename}")

        # Text, PDF oder Bild unterscheiden
        if filename.endswith(".txt"):
            # Textdatei direkt lesen
            print("[UPLOAD] Processing as text file...")
            text_content = contents.decode("utf-8")
            extracted_texts.append(text_content)
            
        elif filename.endswith(".pdf"):
            print("[UPLOAD] Processing as PDF...")
            pdf = fitz.open(stream=BytesIO(contents), filetype="pdf")
            for page_num in range(len(pdf)):
                print(f"[UPLOAD] Processing page {page_num + 1}/{len(pdf)}...")
                page = pdf.load_page(page_num)
                pix = page.get_pixmap(dpi=150)
                img_b64 = base64.b64encode(pix.tobytes("jpeg")).decode("utf-8")

                vision_response = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[
                        {
                            "role": "user",
                            "content": [
                                {
                                    "type": "text",
                                    "text": (
                                        f"Lies den Inhalt dieser Seite ({page_num + 1}) "
                                        "mit allen Mathematikaufgaben und gib NUR den erkannten Text wieder:"
                                    ),
                                },
                                {
                                    "type": "image_url",
                                    "image_url": {
                                        "url": f"data:image/jpeg;base64,{img_b64}"
                                    },
                                },
                            ],
                        }
                    ],
                )
                extracted_texts.append(vision_response.choices[0].message.content)
                
        else:
            # Einzelbild direkt verarbeiten
            print("[UPLOAD] Processing as image...")
            b64 = base64.b64encode(contents).decode("utf-8")
            vision_response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": "Lies den Inhalt dieser Aufgabe und gib NUR den Text wieder:",
                            },
                            {
                                "type": "image_url",
                                "image_url": {"url": f"data:image/jpeg;base64,{b64}"},
                            },
                        ],
                    }
                ],
            )
            extracted_texts.append(vision_response.choices[0].message.content)

        # Gesamttext zusammenfuehren
        print("[UPLOAD] Merging extracted text...")
        full_text = "\n\n".join(extracted_texts)
        print(f"[UPLOAD] Extracted text length: {len(full_text)} characters")

        # Clarity-Coach-Logik auf den erkannten Text anwenden
        print("[UPLOAD] Running Clarity Coach analysis...")
        result = run_clarity_coach(full_text)
        
        print("[UPLOAD] Analysis complete!")
        return result
    
    except Exception as e:
        error_msg = str(e)
        print(f"[ERROR] Upload failed: {error_msg}")
        
        raise HTTPException(
            status_code=500, 
            detail=f"Upload/Analysis failed: {error_msg[:200]}"
        )


# ------------------------------
# Visualization für eine Teilaufgabe generieren
# ------------------------------
@app.post("/visualize")
async def visualize(payload: dict = Body(...)):
    task_number = payload.get("taskNumber")
    task_text = payload.get("taskText", "")
    topic = payload.get("topic", "")
    sub_label = payload.get("subLabel")
    subtask_text = payload.get("subtaskText", "")

    if not subtask_text or not str(subtask_text).strip():
        raise HTTPException(status_code=400, detail="Keine Teilaufgabe übergeben.")

    visualize_prompt = f"""
Du bist ein Mathematik-Experte, der auf die Erstellung klarer, strukturierter Visualisierungen mathematischer Konzepte spezialisiert ist.

Erstelle eine hilfreiche Visualisierung für die folgende Teilaufgabe, die SCHLÜSSELFAKTEN und konzeptionelles Verständnis hervorhebt.

**Aufgabe {task_number}: {topic}**
Hauptaufgabe: {task_text}

**Teilaufgabe {sub_label}:**
{subtask_text}

Erstelle eine strukturierte Visualisierung, die Folgendes enthält:

1. **Kernkonzepte**: Liste der beteiligten mathematischen Konzepte
2. **Gegebene Informationen**: Welche Daten/Informationen sind gegeben
3. **Gesuchtes**: Was soll gelöst oder bewiesen werden
4. **Relevante Formeln**: Wichtige Formeln oder Sätze, die gelten
5. **Wichtige Fakten**: Kritische Fakten oder Eigenschaften zum Merken
6. **Lösungsansatz-Hinweise**: Strategie auf hohem Niveau (ohne die Lösung zu verraten)

FORMATIERE DEINE ANTWORT WIE FOLGT:
- Verwende ** ** für Abschnittsüberschriften (z.B., **Kernkonzepte**)
- Verwende Aufzählungspunkte mit "- " für Listenelemente
- Verwende Schlüssel-Wert-Paare mit ":" für strukturierte Daten
- Füge mathematische Notation mit LaTeX ein, wo angebracht ($...$ für inline, $$...$$ für display)
- Halte Erklärungen prägnant und visuell
- Fokus auf VERSTÄNDNIS, nicht auf die Lösung

Beispielformat:
**Kernkonzepte**
- Polynomfunktionen
- Nullstellen

**Gegebene Informationen**
- Funktion: $f(x) = x^3 - 27$
- Definitionsbereich: Reelle Zahlen

**Gesuchtes**
- Alle reellen Lösungen

Gib NICHT die vollständige Lösung. Konzentriere dich auf die visuelle Organisation der Fakten und Konzepte.
"""

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You are a mathematics visualization assistant."},
                {"role": "user", "content": visualize_prompt}
            ],
            temperature=0.7
        )

        visualization_text = response.choices[0].message.content

        # Safe printing for Windows console
        try:
            print("\n--- Visualization Generated ---\n", visualization_text[:200], "...\n")
        except UnicodeEncodeError:
            print("\n--- Visualization Generated ---\n[Contains Unicode characters]\n")

        return {"visualization": visualization_text}

    except Exception as e:
        print(f"[ERROR] Error generating visualization: {e}")
        raise HTTPException(status_code=500, detail=f"Visualization generation failed: {str(e)}")


# ------------------------------
# Manim Animation für eine Teilaufgabe generieren
# ------------------------------
@app.post("/animate")
async def animate(payload: dict = Body(...)):
    task_number = payload.get("taskNumber")
    task_text = payload.get("taskText", "")
    topic = payload.get("topic", "")
    sub_label = payload.get("subLabel")
    subtask_text = payload.get("subtaskText", "")

    if not subtask_text or not str(subtask_text).strip():
        raise HTTPException(status_code=400, detail="Keine Teilaufgabe übergeben.")

    animate_prompt = f"""
Du bist ein Experte für mathematische Animationen. Erstelle eine Schritt-für-Schritt Animation für folgende Aufgabe.

**Aufgabe {task_number}: {topic}**
Hauptaufgabe: {task_text}

**Teilaufgabe {sub_label}:**
{subtask_text}

Erstelle eine JSON-Struktur für eine Browser-Animation (mit Canvas/SVG + GSAP + KaTeX).

ANFORDERUNGEN:
1. 3-5 klare Animationsschritte
2. Jeder Schritt hat: Text-Erklärung + mathematische Formel (LaTeX)
3. Beschreibe visuelle Effekte (fadeIn, scale, move, highlight, etc.)
4. Halte es einfach und verständlich

ANTWORTFORMAT (reines JSON, keine Erklärungen):
{{
  "title": "Kurzer Titel der Aufgabe",
  "steps": [
    {{
      "id": 1,
      "description": "Beschreibung was passiert",
      "latex": "x^2 + 3x - 4 = 0",
      "animation": "fadeIn",
      "duration": 1.0,
      "position": "center"
    }},
    {{
      "id": 2,
      "description": "Nächster Schritt",
      "latex": "x^2 + 3x = 4",
      "animation": "highlight",
      "duration": 0.8,
      "position": "center",
      "highlight": "+3x"
    }},
    {{
      "id": 3,
      "description": "Transformation",
      "latex": "(x + 4)(x - 1) = 0",
      "animation": "transform",
      "duration": 1.2,
      "position": "center"
    }}
  ]
}}

VERFÜGBARE ANIMATIONEN:
- fadeIn: Element erscheint
- fadeOut: Element verschwindet
- scale: Element wird größer/kleiner
- move: Element bewegt sich
- highlight: Teil wird hervorgehoben
- transform: Sanfte Umwandlung
- bounce: Springender Effekt

POSITIONEN: "center", "top", "bottom", "left", "right"

Gib NUR das JSON zurück, keine Markdown-Codeblöcke, keine Erklärungen.
"""

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "Du bist ein Experte für mathematische Visualisierungen und Animationen."},
                {"role": "user", "content": animate_prompt}
            ],
            temperature=0.7,
            response_format={"type": "json_object"}
        )

        animation_json_str = response.choices[0].message.content
        
        # Safe printing
        try:
            print("\n--- Animation JSON Generated ---\n", animation_json_str[:300], "...\n")
        except UnicodeEncodeError:
            print("\n--- Animation JSON Generated ---\n[Contains Unicode characters]\n")

        # Parse JSON
        animation_data = json.loads(animation_json_str)
        
        # Return the animation data for browser rendering
        return {
            "animationData": animation_data,
            "success": True
        }

    except json.JSONDecodeError as e:
        print(f"[ERROR] JSON Parse Error: {e}")
        raise HTTPException(status_code=500, detail=f"Animation JSON parsing failed: {str(e)}")
    except Exception as e:
        print(f"[ERROR] Error generating animation: {e}")
        raise HTTPException(status_code=500, detail=f"Animation generation failed: {str(e)}")


# ------------------------------
# Plotly Graph für eine Teilaufgabe generieren
# ------------------------------
@app.post("/plot")
async def plot_task(payload: dict = Body(...)):
    task_number = payload.get("taskNumber")
    task_text = payload.get("taskText", "")
    topic = payload.get("topic", "")
    sub_label = payload.get("subLabel")
    subtask_text = payload.get("subtaskText", "")

    if not subtask_text or not str(subtask_text).strip():
        raise HTTPException(status_code=400, detail="Keine Teilaufgabe übergeben.")

    plot_prompt = f"""
Du bist ein Experte für mathematische Visualisierungen. Analysiere die folgende Aufgabe und bestimme, ob eine grafische Darstellung sinnvoll ist.

**Aufgabe {task_number}: {topic}**
Hauptaufgabe: {task_text}

**Teilaufgabe {sub_label}:**
{subtask_text}

ENTSCHEIDE:
1. Ist eine grafische Darstellung für diese Aufgabe möglich und sinnvoll?
2. Hat die Aufgabe eine KONKRETE Funktion mit SPEZIFISCHEN Werten? (Nicht abstrakt mit a, b, c)
3. Wenn JA: Welche Art von Graph? (Funktionsplot, Polynom, Parabel, Gerade, Kreis, etc.)
4. Extrahiere die relevanten Parameter

WICHTIG:
- Wenn die Aufgabe nur abstrakte Bedingungen wie "f'(x₁) = 0" oder "f''(x₁) ≠ 0" enthält OHNE konkrete Funktion → plottable: false
- Wenn die Aufgabe Parameter wie a, b, c enthält OHNE konkrete Werte → plottable: false  
- NUR wenn eine konkrete Funktion wie "x³ - 27" oder "2x + 3" gegeben ist → plottable: true

ANTWORTFORMAT (reines JSON):
{{
  "plottable": true/false,
  "reason": "Kurze Begründung warum ja/nein",
  "graphType": "function" / "polynomial" / "line" / "circle" / "points" / "none",
  "function": "mathematischer Ausdruck in Python-Syntax, z.B. x**3 - 27",
  "domain": {{"xMin": -10, "xMax": 10, "yMin": -50, "yMax": 50}},
  "title": "Titel für die Grafik",
  "xLabel": "x",
  "yLabel": "y oder f(x)",
  "points": [  // Nur wenn graphType = "points"
    {{"x": 1, "y": 2, "label": "Punkt A"}},
    {{"x": 3, "y": 4, "label": "Punkt B"}}
  ],
  "specialPoints": [  // Wichtige Punkte wie Nullstellen, Extrema
    {{"x": 3, "y": 0, "label": "Nullstelle", "color": "red"}}
  ]
}}

BEISPIELE:

Aufgabe: "Löse x^3 - 27 = 0"
{{
  "plottable": true,
  "reason": "Funktionsplot zeigt die Nullstelle visuell",
  "graphType": "polynomial",
  "function": "x**3 - 27",
  "domain": {{"xMin": -5, "xMax": 5, "yMin": -50, "yMax": 50}},
  "title": "f(x) = x³ - 27",
  "xLabel": "x",
  "yLabel": "f(x)",
  "specialPoints": [{{"x": 3, "y": 0, "label": "Nullstelle x=3", "color": "red"}}]
}}

Aufgabe: "Berechne 5 + 3"
{{
  "plottable": false,
  "reason": "Einfache Arithmetik ohne grafische Komponente",
  "graphType": "none"
}}

Aufgabe: "Zeige, dass f'(x₁) = 0 und f''(x₁) ≠ 0 gilt"
{{
  "plottable": false,
  "reason": "Abstrakte Bedingung ohne konkrete Funktion - keine spezifischen Werte für f(x)",
  "graphType": "none"
}}

Aufgabe: "Zeichne die Gerade durch A(1,2) und B(3,6)"
{{
  "plottable": true,
  "reason": "Gerade mit gegebenen Punkten",
  "graphType": "line",
  "function": "2*x",
  "domain": {{"xMin": 0, "xMax": 5, "yMin": 0, "yMax": 10}},
  "title": "Gerade durch A und B",
  "xLabel": "x",
  "yLabel": "y",
  "points": [
    {{"x": 1, "y": 2, "label": "A"}},
    {{"x": 3, "y": 6, "label": "B"}}
  ]
}}

Gib NUR das JSON zurück, keine Erklärungen, keine Markdown-Codeblöcke.
"""

    try:
        print("[PLOT] Analyzing task for plottability...")
        
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "Du bist ein Experte für mathematische Visualisierungen. Antworte nur mit gültigem JSON."},
                {"role": "user", "content": plot_prompt}
            ],
            temperature=0.7,
            response_format={"type": "json_object"}
        )

        plot_json_str = response.choices[0].message.content
        print(f"[PLOT] GPT Response: {plot_json_str[:200]}...")
        
        # Parse JSON
        plot_data = json.loads(plot_json_str)
        
        # Check if plottable
        if not plot_data.get("plottable", False):
            return {
                "plottable": False,
                "message": plot_data.get("reason", "Keine grafische Darstellung möglich für diese Aufgabe")
            }
        
        # Generate plot based on graphType - using Chart.js format (Phase 2.2)
        graph_type = plot_data.get("graphType", "function")
        
        print(f"[PLOT] Generating {graph_type} plot for Chart.js...")
        
        if graph_type in ["function", "polynomial", "line"]:
            # Function plot - generate data points for Chart.js
            func_str = plot_data.get("function", "x")
            domain = plot_data.get("domain", {"xMin": -10, "xMax": 10, "yMin": -50, "yMax": 50})
            
            # Generate x values (100 points for smooth curve, less than Plotly's 500)
            x_values = np.linspace(domain.get("xMin", -10), domain.get("xMax", 10), 100)
            
            # Evaluate function safely
            try:
                # Replace common math notation
                func_str_safe = func_str.replace("^", "**").replace("π", "np.pi").replace("e", "np.e")
                
                # Check if function has undefined variables
                test_x = np.array([1.0])
                test_eval = eval(func_str_safe, {"x": test_x, "np": np, "__builtins__": {}})
                
                # Evaluate for full domain
                y_values = eval(func_str_safe, {"x": x_values, "np": np, "__builtins__": {}})
            except (NameError, SyntaxError) as e:
                print(f"[ERROR] Function contains undefined variables or syntax error: {e}")
                return {
                    "plottable": False,
                    "message": "Diese Aufgabe ist zu abstrakt für eine konkrete Grafik. Sie enthält allgemeine Parameter ohne spezifische Werte."
                }
            except Exception as e:
                print(f"[ERROR] Function evaluation failed: {e}")
                return {
                    "plottable": False,
                    "message": f"Konnte Funktion nicht auswerten: {str(e)}"
                }
            
            # Convert numpy arrays to lists and create Chart.js data format
            # Create xy pairs for scatter chart (line type)
            data_points = [{"x": float(x), "y": float(y)} for x, y in zip(x_values, y_values) if np.isfinite(y)]
            
            # Prepare special points (nullstellen, extrema, etc.)
            special_points_data = []
            for point in plot_data.get("specialPoints", []):
                special_points_data.append({
                    "x": point.get("x"),
                    "y": point.get("y"),
                    "label": point.get("label", ""),
                    "color": point.get("color", "#ef4444")  # Default red
                })
            
            # Prepare given points
            points_data = []
            for point in plot_data.get("points", []):
                points_data.append({
                    "x": point.get("x"),
                    "y": point.get("y"),
                    "label": point.get("label", ""),
                    "color": "#10b981"  # Green
                })
            
            # Chart.js configuration
            chart_config = {
                "type": "scatter",
                "data": {
                    "datasets": [
                        {
                            "label": plot_data.get("title", "f(x)"),
                            "data": data_points,
                            "borderColor": "#2c5f8d",
                            "backgroundColor": "rgba(44, 95, 141, 0.1)",
                            "borderWidth": 2,
                            "pointRadius": 0,
                            "showLine": True,
                            "tension": 0.1,
                            "fill": False
                        }
                    ]
                },
                "options": {
                    "responsive": True,
                    "maintainAspectRatio": False,
                    "plugins": {
                        "title": {
                            "display": True,
                            "text": plot_data.get("title", "Grafik"),
                            "font": {"size": 16, "weight": "bold"}
                        },
                        "legend": {
                            "display": True,
                            "position": "top"
                        },
                        "tooltip": {
                            "enabled": True,
                            "mode": "nearest",
                            "intersect": False
                        }
                    },
                    "scales": {
                        "x": {
                            "type": "linear",
                            "position": "center",
                            "title": {
                                "display": True,
                                "text": plot_data.get("xLabel", "x")
                            },
                            "min": domain.get("xMin", -10),
                            "max": domain.get("xMax", 10),
                            "grid": {
                                "color": "rgba(0, 0, 0, 0.1)"
                            }
                        },
                        "y": {
                            "type": "linear",
                            "position": "center",
                            "title": {
                                "display": True,
                                "text": plot_data.get("yLabel", "y")
                            },
                            "min": domain.get("yMin", -50),
                            "max": domain.get("yMax", 50),
                            "grid": {
                                "color": "rgba(0, 0, 0, 0.1)"
                            }
                        }
                    },
                    "interaction": {
                        "mode": "nearest",
                        "axis": "x",
                        "intersect": False
                    }
                },
                "specialPoints": special_points_data,
                "givenPoints": points_data
            }
            
            # Add special points as separate datasets
            if special_points_data:
                chart_config["data"]["datasets"].append({
                    "label": "Besondere Punkte",
                    "data": [{"x": p["x"], "y": p["y"]} for p in special_points_data],
                    "borderColor": "#ef4444",
                    "backgroundColor": "#ef4444",
                    "pointRadius": 8,
                    "pointHoverRadius": 10,
                    "showLine": False
                })
            
            if points_data:
                chart_config["data"]["datasets"].append({
                    "label": "Gegebene Punkte",
                    "data": [{"x": p["x"], "y": p["y"]} for p in points_data],
                    "borderColor": "#10b981",
                    "backgroundColor": "#10b981",
                    "pointRadius": 8,
                    "pointHoverRadius": 10,
                    "showLine": False
                })
            
            print("[PLOT] Chart.js data generated successfully!")
            
            return {
                "plotData": json.dumps(chart_config),
                "plottable": True,
                "graphType": graph_type,
                "chartType": "chartjs"  # Indicator for frontend
            }
        
        elif graph_type == "points":
            # Scatter plot with points only
            points = plot_data.get("points", [])
            
            if not points:
                return {
                    "plottable": False,
                    "message": "Keine Punkte zum Plotten vorhanden"
                }
            
            domain = plot_data.get("domain", {"xMin": -10, "xMax": 10, "yMin": -10, "yMax": 10})
            
            # Prepare points data
            points_data = [{"x": p.get("x"), "y": p.get("y")} for p in points]
            point_labels = [p.get("label", f"P{i+1}") for i, p in enumerate(points)]
            
            chart_config = {
                "type": "scatter",
                "data": {
                    "datasets": [{
                        "label": plot_data.get("title", "Punkte"),
                        "data": points_data,
                        "borderColor": "#2c5f8d",
                        "backgroundColor": "#2c5f8d",
                        "pointRadius": 8,
                        "pointHoverRadius": 10,
                        "showLine": False
                    }]
                },
                "options": {
                    "responsive": True,
                    "maintainAspectRatio": False,
                    "plugins": {
                        "title": {
                            "display": True,
                            "text": plot_data.get("title", "Punkte"),
                            "font": {"size": 16, "weight": "bold"}
                        },
                        "tooltip": {
                            "enabled": True,
                            "callbacks": {}  # Will use default
                        }
                    },
                    "scales": {
                        "x": {
                            "type": "linear",
                            "position": "center",
                            "title": {
                                "display": True,
                                "text": plot_data.get("xLabel", "x")
                            },
                            "min": domain.get("xMin", -10),
                            "max": domain.get("xMax", 10),
                            "grid": {"color": "rgba(0, 0, 0, 0.1)"}
                        },
                        "y": {
                            "type": "linear", 
                            "position": "center",
                            "title": {
                                "display": True,
                                "text": plot_data.get("yLabel", "y")
                            },
                            "min": domain.get("yMin", -10),
                            "max": domain.get("yMax", 10),
                            "grid": {"color": "rgba(0, 0, 0, 0.1)"}
                        }
                    }
                },
                "pointLabels": point_labels
            }
            
            return {
                "plotData": json.dumps(chart_config),
                "plottable": True,
                "graphType": graph_type,
                "chartType": "chartjs"
            }
        
        else:
            return {
                "plottable": False,
                "message": f"Graph-Typ '{graph_type}' noch nicht implementiert"
            }

    except json.JSONDecodeError as e:
        print(f"[ERROR] JSON Parse Error: {e}")
        raise HTTPException(status_code=500, detail=f"Plot JSON parsing failed: {str(e)}")
    except Exception as e:
        print(f"[ERROR] Error generating plot: {e}")
        raise HTTPException(status_code=500, detail=f"Plot generation failed: {str(e)}")


# ------------------------------
# Session Logging - Excel Integration
# ------------------------------
class SessionLogEntry(BaseModel):
    benutzer_name: str
    klasse: str
    schule: str
    fach: str
    thema: str
    aufgabentyp: str
    schwierigkeitsgrad: str
    datei_name: str
    datei_typ: str
    anzahl_aufgaben: int
    anzahl_teilaufgaben: int
    visualisierungen_genutzt: int
    animationen_genutzt: int
    grafiken_genutzt: int
    hints_genutzt: int  # Replaced loesungen_angezeigt (solutions violated Socratic method)
    ansatzpruefungen_genutzt: int = 0  # Phase 3.3: Approach checks used
    selbststaendigkeits_score: int = 5  # Phase 3.3: Self-sufficiency (1-5)
    feedback: str
    sitzungsdauer_minuten: float
    notizen: str


EXCEL_PATH = r"C:\Users\admin\Desktop\Sonstiges\HMS_PROJEKT\clarity-coach\Clarity_Coach_Session_Log.xlsx"


@app.post("/log-session")
async def log_session(entry: SessionLogEntry):
    """
    Log a Clarity Coach session to Excel file
    """
    try:
        print(f"[LOG] Starting session log...")
        
        # Validate required fields
        if not entry.benutzer_name or not entry.benutzer_name.strip():
            raise HTTPException(status_code=400, detail="benutzer_name is required")
        if not entry.klasse or not entry.klasse.strip():
            raise HTTPException(status_code=400, detail="klasse is required")
        if not entry.schule or not entry.schule.strip():
            raise HTTPException(status_code=400, detail="schule is required")
        if not entry.fach or not entry.fach.strip():
            raise HTTPException(status_code=400, detail="fach is required")
        if not entry.thema or not entry.thema.strip():
            raise HTTPException(status_code=400, detail="thema is required")
        if not entry.schwierigkeitsgrad or not entry.schwierigkeitsgrad.strip():
            raise HTTPException(status_code=400, detail="schwierigkeitsgrad is required")
        
        # Validate numeric fields
        if entry.anzahl_aufgaben < 0:
            raise HTTPException(status_code=400, detail="anzahl_aufgaben must be non-negative")
        if entry.anzahl_teilaufgaben < 0:
            raise HTTPException(status_code=400, detail="anzahl_teilaufgaben must be non-negative")
        if entry.visualisierungen_genutzt < 0:
            raise HTTPException(status_code=400, detail="visualisierungen_genutzt must be non-negative")
        if entry.animationen_genutzt < 0:
            raise HTTPException(status_code=400, detail="animationen_genutzt must be non-negative")
        if entry.grafiken_genutzt < 0:
            raise HTTPException(status_code=400, detail="grafiken_genutzt must be non-negative")
        if entry.hints_genutzt < 0:
            raise HTTPException(status_code=400, detail="hints_genutzt must be non-negative")
        if entry.ansatzpruefungen_genutzt < 0:
            raise HTTPException(status_code=400, detail="ansatzpruefungen_genutzt must be non-negative")
        if entry.selbststaendigkeits_score < 1 or entry.selbststaendigkeits_score > 5:
            raise HTTPException(status_code=400, detail="selbststaendigkeits_score must be between 1 and 5")
        if entry.sitzungsdauer_minuten < 0:
            raise HTTPException(status_code=400, detail="sitzungsdauer_minuten must be non-negative")
        
        # Check if Excel file exists
        if not os.path.exists(EXCEL_PATH):
            print(f"[ERROR] Excel file not found at: {EXCEL_PATH}")
            raise HTTPException(
                status_code=500, 
                detail=f"Excel file not found. Please create it first using create_excel_template.py"
            )
        
        # Load workbook
        wb = load_workbook(EXCEL_PATH)
        
        # Get or create Session_Log sheet
        if "Session_Log" not in wb.sheetnames:
            print(f"[ERROR] Session_Log sheet not found")
            raise HTTPException(
                status_code=500,
                detail="Session_Log sheet not found in Excel file"
            )
        
        ws = wb["Session_Log"]
        
        # Update headers if they're old format (with underscores)
        # This ensures existing files get updated headers without losing data
        if ws.max_row > 0:
            header_row = 1
            old_headers = [
                "Session_ID", "Benutzer_Name", "Datei_Name", "Datei_Typ",
                "Anzahl_Aufgaben", "Anzahl_Teilaufgaben", "Visualisierungen_Genutzt",
                "Animationen_Genutzt", "Grafiken_Genutzt", "Hints_Genutzt",
                "Ansatzpruefungen_Genutzt", "Selbststaendigkeits_Score",
                "Sitzungsdauer_Minuten"
            ]
            new_headers = [
                "Session-ID", "Benutzer Name", "Datei Name", "Datei Typ",
                "Anzahl Aufgaben", "Anzahl Teilaufgaben", "Visualisierungen Genutzt",
                "Animationen Genutzt", "Grafiken Genutzt", "Hilfestellungen Genutzt",
                "Ansatzprüfungen Genutzt", "Selbstständigkeits Score",
                "Sitzungsdauer (Minuten)"
            ]
            
            # Check if headers need updating
            needs_update = False
            for col in range(1, min(ws.max_column + 1, 24)):
                cell_value = ws.cell(row=header_row, column=col).value
                if cell_value and any(old in str(cell_value) for old in old_headers):
                    needs_update = True
                    break
            
            if needs_update:
                # Update headers to proper German names
                proper_headers = [
                    "Session-ID", "Datum", "Uhrzeit", "Benutzer Name", "Klasse", "Schule",
                    "Fach", "Thema", "Aufgabentyp", "Schwierigkeitsgrad", "Datei Name",
                    "Datei Typ", "Anzahl Aufgaben", "Anzahl Teilaufgaben",
                    "Visualisierungen Genutzt", "Animationen Genutzt", "Grafiken Genutzt",
                    "Hilfestellungen Genutzt", "Ansatzprüfungen Genutzt",
                    "Selbstständigkeits Score", "Feedback", "Sitzungsdauer (Minuten)", "Notizen"
                ]
                for col, header in enumerate(proper_headers, 1):
                    ws.cell(row=header_row, column=col).value = header
                print("[LOG] Updated Session_Log headers to proper German names")
        
        # Find next row
        next_row = ws.max_row + 1
        
        # Generate Session ID (format: YYYYMMDD-###)
        today = datetime.now().strftime("%Y%m%d")
        session_count = next_row - 1  # -1 for header
        session_id = f"{today}-{session_count:03d}"
        
        # Get current date and time (DD.MM.YYYY format)
        current_date = datetime.now().strftime("%d.%m.%Y")
        current_time = datetime.now().strftime("%H:%M:%S")
        
        # Prepare data row
        data_row = [
            session_id,
            current_date,
            current_time,
            entry.benutzer_name,
            entry.klasse,
            entry.schule,
            entry.fach,
            entry.thema,
            entry.aufgabentyp,
            entry.schwierigkeitsgrad,
            entry.datei_name,
            entry.datei_typ,
            entry.anzahl_aufgaben,
            entry.anzahl_teilaufgaben,
            entry.visualisierungen_genutzt,
            entry.animationen_genutzt,
            entry.grafiken_genutzt,
            entry.hints_genutzt,  # Replaced loesungen_angezeigt
            entry.ansatzpruefungen_genutzt,  # Phase 3.3: Approach checks
            entry.selbststaendigkeits_score,  # Phase 3.3: Self-sufficiency score (1-5)
            entry.feedback,
            entry.sitzungsdauer_minuten,
            entry.notizen
        ]
        
        # Write data to row
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, value in enumerate(data_row, 1):
            cell = ws.cell(row=next_row, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Save workbook
        wb.save(EXCEL_PATH)
        wb.close()
        
        print(f"[LOG] Session logged successfully: {session_id}")
        
        return {
            "success": True,
            "session_id": session_id,
            "message": "Session erfolgreich protokolliert",
            "row": next_row
        }
    
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"[ERROR] Failed to log session: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to log session: {str(e)}"
        )

# ------------------------------
# Get Assessment Data for Dashboard
# ------------------------------
@app.get("/get-assessments")
async def get_assessments():
    """
    Get all assessment data for dashboard analytics.
    Returns assessment data from Assessment_Log sheet.
    """
    try:
        if not os.path.exists(EXCEL_PATH):
            return {"assessments": []}
        
        wb = load_workbook(EXCEL_PATH)
        
        if "Assessment_Log" not in wb.sheetnames:
            wb.close()
            return {"assessments": []}
        
        ws = wb["Assessment_Log"]
        assessments = []
        
        # Read headers
        headers = []
        for col in range(1, ws.max_column + 1):
            headers.append(ws.cell(row=1, column=col).value or "")
        
        # Read data rows
        for row in range(2, ws.max_row + 1):
            assessment = {}
            for col, header in enumerate(headers, 1):
                value = ws.cell(row=row, column=col).value
                # Convert header to camelCase for frontend
                key = header.lower().replace(" ", "_").replace("(", "").replace(")", "").replace("-", "_")
                assessment[key] = value
            assessments.append(assessment)
        
        wb.close()
        
        return {"assessments": assessments}
    
    except Exception as e:
        print(f"[ERROR] Failed to get assessments: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get assessments: {str(e)}")

# ------------------------------
# Assessment Logging Endpoint
# ------------------------------
@app.post("/log-assessment")
async def log_assessment(assessment_data: dict = Body(...)):
    """
    Log post-session assessment to separate Excel sheet (Assessment_Log).
    This endpoint receives tutor/evaluator ratings and observations.
    """
    try:
        print("[ASSESSMENT] Starting assessment log...")
        
        # Validate required fields
        required_fields = ["sessionId", "studentId", "aiQuestionQuality", "engagementLevel", "understandingProgress", "efficiencyScore"]
        for field in required_fields:
            if field not in assessment_data or assessment_data[field] is None or assessment_data[field] == "":
                raise HTTPException(status_code=400, detail=f"Required field missing: {field}")
        
        # Validate scale fields (1-5)
        scale_fields = {
            "aiQuestionQuality": assessment_data.get("aiQuestionQuality"),
            "engagementLevel": assessment_data.get("engagementLevel"),
            "understandingProgress": assessment_data.get("understandingProgress"),
            "efficiencyScore": assessment_data.get("efficiencyScore")
        }
        for field_name, value in scale_fields.items():
            if value is not None:
                try:
                    num_value = int(value)
                    if num_value < 1 or num_value > 5:
                        raise HTTPException(status_code=400, detail=f"{field_name} must be between 1 and 5")
                except (ValueError, TypeError):
                    raise HTTPException(status_code=400, detail=f"{field_name} must be a number between 1 and 5")
        
        # Validate tutorInterventions (must be non-negative integer)
        tutor_interventions = assessment_data.get("tutorInterventions", 0)
        try:
            tutor_interventions = int(tutor_interventions)
            if tutor_interventions < 0:
                raise HTTPException(status_code=400, detail="tutorInterventions must be non-negative")
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="tutorInterventions must be a non-negative integer")
        
        # Check for duplicate session assessment
        if os.path.exists(EXCEL_PATH):
            wb_check = load_workbook(EXCEL_PATH)
            if "Assessment_Log" in wb_check.sheetnames:
                ws_check = wb_check["Assessment_Log"]
                session_id = assessment_data.get("sessionId", "")
                # Check if this session already has an assessment
                for row in range(2, ws_check.max_row + 1):
                    if ws_check.cell(row=row, column=1).value == session_id:
                        wb_check.close()
                        raise HTTPException(status_code=409, detail=f"Assessment for session {session_id} already exists")
            wb_check.close()
        
        # Check if Excel file exists
        if not os.path.exists(EXCEL_PATH):
            raise HTTPException(
                status_code=404,
                detail="Session log file not found. Please complete a session first."
            )
        
        # Load workbook
        wb = load_workbook(EXCEL_PATH)
        
        # Import styling classes OUTSIDE conditional block (needed for data rows too)
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Create Assessment_Log sheet if it doesn't exist
        if "Assessment_Log" not in wb.sheetnames:
            ws = wb.create_sheet("Assessment_Log")
            
            # Define headers (21 columns as per requirements)
            headers = [
                "Session-ID",
                "Date (DD.MM.YYYY)",
                "Student-ID",
                "Grade",
                "Topic Area",
                "Topic Detail",
                "Topic Complexity (1-5)",
                "AI Question Quality (1-5)",
                "Prompt Strategy",
                "Tutor Interventions",
                "Student Override (Yes/No)",
                "Learner Type Indicator",
                "Understanding Progress (1-5)",
                "Linguistic Neutrality Check (Yes/No)",
                "Engagement Level (1-5)",
                "Evaluative Language Check (Yes/No)",
                "Student Feedback Safety (Yes/No/Unclear)",
                "Question Loops",
                "Efficiency Score (1-5)",
                "Remarks",
                "Further Considerations"
            ]
            
            # Write headers
            ws.append(headers)
            
            # Style headers (matching Session_Log style)
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(
                start_color="2C5F8D",
                end_color="2C5F8D",
                fill_type="solid"
            )
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                cell = ws[f"{col_letter}1"]
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                
                # Set column widths
                if header in ["Remarks", "Further Considerations", "Prompt Strategy"]:
                    ws.column_dimensions[col_letter].width = 40
                elif header == "Session-ID":
                    ws.column_dimensions[col_letter].width = 15
                elif header in ["Learner Type Indicator", "Topic Area", "Topic Detail"]:
                    ws.column_dimensions[col_letter].width = 25
                elif header == "Date (DD.MM.YYYY)":
                    ws.column_dimensions[col_letter].width = 16
                else:
                    ws.column_dimensions[col_letter].width = 18
            
            print("[ASSESSMENT] Created new Assessment_Log sheet")
        else:
            ws = wb["Assessment_Log"]
        
        # Get topic data from Session_Log by joining on Session-ID
        topic_area = ""
        topic_detail = ""
        topic_complexity = ""
        session_id = assessment_data.get("sessionId", "")
        
        if "Session_Log" in wb.sheetnames:
            session_ws = wb["Session_Log"]
            # Find session in Session_Log (Session_ID is column A, Thema is column H, Schwierigkeitsgrad is column J)
            for row in range(2, session_ws.max_row + 1):
                if session_ws.cell(row=row, column=1).value == session_id:
                    topic_area = session_ws.cell(row=row, column=7).value or ""  # Fach
                    topic_detail = session_ws.cell(row=row, column=8).value or ""  # Thema
                    difficulty = session_ws.cell(row=row, column=10).value or ""  # Schwierigkeitsgrad
                    # Convert difficulty to 1-5 scale
                    if difficulty == "Leicht":
                        topic_complexity = 1
                    elif difficulty == "Mittel":
                        topic_complexity = 3
                    elif difficulty == "Anspruchsvoll":
                        topic_complexity = 5
                    else:
                        topic_complexity = ""
                    break
        
        # Prepare assessment data row (21 columns)
        now = datetime.now()
        # Format date as DD.MM.YYYY
        date_str = now.strftime("%d.%m.%Y")
        
        # Convert boolean to Yes/No
        student_override_str = "Yes" if assessment_data.get("studentOverride", False) else "No"
        linguistic_neutrality_str = "Yes" if assessment_data.get("linguisticNeutralityCheck", False) else "No"
        evaluative_language_str = "Yes" if assessment_data.get("evaluativeLanguageCheck", False) else "No"
        student_feedback_safety = assessment_data.get("studentFeedbackSafety", "") or ""
        if student_feedback_safety:
            student_feedback_safety = student_feedback_safety.capitalize()
        
        assessment_row = [
            session_id,
            date_str,  # DD.MM.YYYY format
            assessment_data.get("studentId", ""),
            assessment_data.get("grade", ""),
            topic_area,  # From Session_Log
            topic_detail,  # From Session_Log
            topic_complexity,  # From Session_Log (1-5)
            assessment_data.get("aiQuestionQuality", 0),
            assessment_data.get("promptStrategy", ""),
            assessment_data.get("tutorInterventions", 0),
            student_override_str,
            assessment_data.get("learnerTypeIndicator", "Nicht angegeben"),
            assessment_data.get("understandingProgress", 0),
            linguistic_neutrality_str,
            assessment_data.get("engagementLevel", 0),
            evaluative_language_str,
            student_feedback_safety,
            assessment_data.get("questionLoops", 0),
            assessment_data.get("efficiencyScore", 0),
            assessment_data.get("remarks", ""),
            assessment_data.get("furtherConsiderations", "")
        ]
        
        # Find next row
        next_row = ws.max_row + 1
        
        # Write data with formatting
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, value in enumerate(assessment_row, 1):
            cell = ws.cell(row=next_row, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Save workbook
        wb.save(EXCEL_PATH)
        wb.close()
        
        session_id = assessment_data.get("sessionId", "Unknown")
        print(f"[ASSESSMENT] Assessment logged successfully for session: {session_id}")
        
        return {
            "success": True,
            "sessionId": session_id,
            "message": "Assessment erfolgreich protokolliert",
            "row": next_row
        }
    
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"[ERROR] Failed to log assessment: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Failed to log assessment: {str(e)}"
        )

# ------------------------------
# Export Assessment Log to Separate Excel File
# ------------------------------
@app.get("/export-assessment-log")
async def export_assessment_log():
    """
    Export all assessment data to a separate Excel file.
    File name: Clarity_Coach_Assessment_Log_[DATE].xlsx
    Contains exactly 21 columns as per Assessment template.
    """
    try:
        print("[EXPORT] Starting assessment log export...")
        
        # Check if main Excel file exists
        if not os.path.exists(EXCEL_PATH):
            raise HTTPException(
                status_code=404,
                detail="Session log file not found. No assessments to export."
            )
        
        # Load workbook
        wb = load_workbook(EXCEL_PATH)
        
        # Check if Assessment_Log sheet exists
        if "Assessment_Log" not in wb.sheetnames:
            wb.close()
            raise HTTPException(
                status_code=404,
                detail="No assessment data found. Please complete at least one assessment first."
            )
        
        # Import styling classes
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Create new workbook for export
        export_wb = Workbook()
        export_ws = export_wb.active
        export_ws.title = "Assessment_Log"
        
        # Get data from existing Assessment_Log
        source_ws = wb["Assessment_Log"]
        
        # Copy headers (21 columns)
        headers = [
            "Session-ID",
            "Date (DD.MM.YYYY)",
            "Student-ID",
            "Grade",
            "Topic Area",
            "Topic Detail",
            "Topic Complexity (1-5)",
            "AI Question Quality (1-5)",
            "Prompt Strategy",
            "Tutor Interventions",
            "Student Override (Yes/No)",
            "Learner Type Indicator",
            "Understanding Progress (1-5)",
            "Linguistic Neutrality Check (Yes/No)",
            "Engagement Level (1-5)",
            "Evaluative Language Check (Yes/No)",
            "Student Feedback Safety (Yes/No/Unclear)",
            "Question Loops",
            "Efficiency Score (1-5)",
            "Remarks",
            "Further Considerations"
        ]
        
        # Write headers
        export_ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="2C5F8D",
            end_color="2C5F8D",
            fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = export_ws[f"{col_letter}1"]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            
            # Set column widths
            if header in ["Remarks", "Further Considerations", "Prompt Strategy"]:
                export_ws.column_dimensions[col_letter].width = 40
            elif header == "Session-ID":
                export_ws.column_dimensions[col_letter].width = 15
            elif header in ["Learner Type Indicator", "Topic Area", "Topic Detail"]:
                export_ws.column_dimensions[col_letter].width = 25
            elif header == "Date (DD.MM.YYYY)":
                export_ws.column_dimensions[col_letter].width = 16
            else:
                export_ws.column_dimensions[col_letter].width = 18
        
        # Copy all data rows (skip header row)
        for row in range(2, source_ws.max_row + 1):
            data_row = []
            for col in range(1, 22):  # Exactly 21 columns
                cell_value = source_ws.cell(row=row, column=col).value
                data_row.append(cell_value)
            
            # Write data row
            export_ws.append(data_row)
            
            # Apply formatting
            for col in range(1, 22):
                cell = export_ws.cell(row=export_ws.max_row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Set row height for header
        export_ws.row_dimensions[1].height = 30
        
        # Freeze first row
        export_ws.freeze_panes = 'A2'
        
        # Generate filename with date
        today = datetime.now().strftime("%Y%m%d")
        export_filename = f"Clarity_Coach_Assessment_Log_{today}.xlsx"
        export_path = os.path.join(
            os.path.dirname(EXCEL_PATH),
            export_filename
        )
        
        # Save export file
        export_wb.save(export_path)
        export_wb.close()
        wb.close()
        
        print(f"[EXPORT] Assessment log exported successfully: {export_filename}")
        
        return {
            "success": True,
            "filename": export_filename,
            "path": export_path,
            "row_count": source_ws.max_row - 1,  # Exclude header
            "message": f"Assessment log exported successfully. {source_ws.max_row - 1} assessments exported."
        }
    
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"[ERROR] Failed to export assessment log: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Failed to export assessment log: {str(e)}"
        )
