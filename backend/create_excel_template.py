"""
Create Excel Session Log Template
This script creates the Clarity_Coach_Session_Log.xlsx file with proper structure
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_session_log_template():
    """Create the Excel template for session logging"""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Session_Log"
    
    # Define headers (Updated Phase 3.3 - Added approach checks and self-sufficiency)
    # All 23 columns with proper German names
    headers = [
        "Session-ID",                    # Column 1
        "Datum",                         # Column 2
        "Uhrzeit",                       # Column 3
        "Benutzer Name",                 # Column 4
        "Klasse",                        # Column 5
        "Schule",                        # Column 6
        "Fach",                          # Column 7
        "Thema",                         # Column 8
        "Aufgabentyp",                   # Column 9
        "Schwierigkeitsgrad",            # Column 10
        "Datei Name",                    # Column 11
        "Datei Typ",                     # Column 12
        "Anzahl Aufgaben",              # Column 13
        "Anzahl Teilaufgaben",          # Column 14
        "Visualisierungen Genutzt",     # Column 15
        "Animationen Genutzt",          # Column 16
        "Grafiken Genutzt",             # Column 17
        "Hilfestellungen Genutzt",      # Column 18 (was Hints_Genutzt)
        "Ansatzprüfungen Genutzt",      # Column 19 (was Ansatzpruefungen_Genutzt)
        "Selbstständigkeits Score",     # Column 20 (was Selbststaendigkeits_Score)
        "Feedback",                      # Column 21
        "Sitzungsdauer (Minuten)",      # Column 22 (was Sitzungsdauer_Minuten)
        "Notizen"                        # Column 23
    ]
    
    # Style definitions
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Set column widths (Updated Phase 3.3 - All 23 columns)
    column_widths = {
        'A': 12,  # Session-ID
        'B': 12,  # Datum
        'C': 10,  # Uhrzeit
        'D': 18,  # Benutzer Name
        'E': 10,  # Klasse
        'F': 20,  # Schule
        'G': 15,  # Fach
        'H': 25,  # Thema
        'I': 18,  # Aufgabentyp
        'J': 16,  # Schwierigkeitsgrad
        'K': 20,  # Datei Name
        'L': 12,  # Datei Typ
        'M': 15,  # Anzahl Aufgaben
        'N': 18,  # Anzahl Teilaufgaben
        'O': 20,  # Visualisierungen Genutzt
        'P': 20,  # Animationen Genutzt
        'Q': 18,  # Grafiken Genutzt
        'R': 20,  # Hilfestellungen Genutzt
        'S': 22,  # Ansatzprüfungen Genutzt
        'T': 22,  # Selbstständigkeits Score
        'U': 15,  # Feedback
        'V': 22,  # Sitzungsdauer (Minuten) - Column 22
        'W': 30,  # Notizen - Column 23
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Add example row (optional, can be removed) - Updated Phase 3.3
    # All 23 columns of example data
    example_data = [
        "001",                                    # Session-ID
        datetime.now().strftime("%d.%m.%Y"),      # Datum (DD.MM.YYYY)
        datetime.now().strftime("%H:%M:%S"),      # Uhrzeit
        "Max Mustermann",                         # Benutzer Name
        "10a",                                    # Klasse
        "Gymnasium Beispiel",                     # Schule
        "Mathematik",                             # Fach
        "Quadratische Gleichungen",               # Thema
        "Gleichungen lösen",                      # Aufgabentyp
        "Mittel",                                 # Schwierigkeitsgrad
        "aufgabe_quadratisch.pdf",                # Datei Name
        "PDF",                                    # Datei Typ
        "1",                                      # Anzahl Aufgaben
        "3",                                      # Anzahl Teilaufgaben
        "2",                                      # Visualisierungen Genutzt
        "1",                                      # Animationen Genutzt
        "0",                                      # Grafiken Genutzt
        "2",                                      # Hilfestellungen Genutzt
        "1",                                      # Ansatzprüfungen Genutzt
        "4",                                      # Selbstständigkeits Score (1-5, 5=independent)
        "Helpful",                                # Feedback
        "15.5",                                   # Sitzungsdauer (Minuten) - Column 22
        "Gute Fortschritte beim Verständnis"      # Notizen - Column 23
    ]
    
    for col_num, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    # Save file
    output_path = r"C:\Users\admin\Desktop\Sonstiges\HMS_PROJEKT\clarity-coach\Clarity_Coach_Session_Log.xlsx"
    wb.save(output_path)
    print(f"[SUCCESS] Excel template created at: {output_path}")
    print(f"[INFO] Template includes {len(headers)} columns")
    return output_path

if __name__ == "__main__":
    create_session_log_template()
